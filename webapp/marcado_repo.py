# ============================================================================
# PROPIEDAD INTELECTUAL Y LICENCIA COMERCIAL CERRADA
# ============================================================================
# Autor Legal y Titular de Derechos: JAVIER ILLAN GONZALEZ
# Organización: ORANGE CREW
# Contacto: ILLANJAVIER9@GMAIL.COM
#
# ADVERTENCIA LEGAL (MÉXICO Y GLOBAL):
# Este código fuente y su arquitectura son propiedad intelectual exclusiva de
# JAVIER ILLAN GONZALEZ. Queda estrictamente prohibida su reproducción,
# distribución, modificación, ingeniería inversa, copia o uso comercial sin la
# autorización expresa y por escrito del autor. Obra protegida conforme a la
# Ley Federal del Derecho de Autor y tratados internacionales aplicables.
# ============================================================================
import os
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from db import get_connection

AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NEGRITA = Font(bold=True)


def generar_marcado(documento_id):
    con = get_connection()
    documento = con.execute(
        "SELECT * FROM documentos_importados WHERE id = ?", (documento_id,)
    ).fetchone()
    if documento is None or not documento["ruta_archivo"]:
        con.close()
        raise ValueError("No se encontrÃ³ el archivo original de este documento.")

    movimientos = con.execute(
        """SELECT m.fila_original, m.fecha, p.tipo, p.numero
           FROM movimientos m
           JOIN polizas p ON p.id = m.poliza_id
           WHERE m.documento_id = ? AND m.fila_original IS NOT NULL""",
        (documento_id,),
    ).fetchall()
    con.close()

    wb = load_workbook(documento["ruta_archivo"])
    hoja = documento["nombre_hoja"] or wb.sheetnames[0]
    ws = wb[hoja]

    ultima_col = ws.max_column
    ws.cell(row=1, column=ultima_col + 1, value="NÃºmero de PÃ³liza").font = NEGRITA

    marcados = 0
    for m in movimientos:
        # fila_original es Ã­ndice 0-based de pandas; +2 por encabezado y base 1 de Excel
        fila_excel = m["fila_original"] + 2
        if fila_excel > ws.max_row:
            continue
        for col in range(1, ultima_col + 1):
            ws.cell(row=fila_excel, column=col).fill = AMARILLO

        etiqueta = f"{'EG' if m['tipo'] == 'egreso' else 'IN'}-{m['numero']}"
        celda = ws.cell(row=fila_excel, column=ultima_col + 1, value=etiqueta)
        celda.font = NEGRITA
        celda.fill = AMARILLO
        marcados += 1

    ruta_salida = os.path.join(
        tempfile.gettempdir(), f"marcado_{os.path.basename(documento['nombre_archivo'])}"
    )
    wb.save(ruta_salida)
    return ruta_salida, marcados

