# -*- coding: utf-8 -*-
"""
Exportador de pólizas al formato de importación masiva de Contpaqi — versión
enriquecida con asociación de CFDI y detalle de impuestos (F4).

Se basa en un archivo de referencia real (proporcionado por el usuario)
que ya traía pólizas con asociación de CFDI y de impuestos. La
estructura real de Contpaqi tiene, además de P y M1:

  AM  Asociación movimiento   -> liga una línea M1 al UUID de un CFDI
  I   Movimientos de impuestos -> el detalle de impuestos (esto es el "F4"):
      base, importe de IVA, tasa, folio y UUID de la factura
  W2  Devolución de IVA (IETU) -> acompaña a cada línea I
  V   Devolución de IVA        -> resumen de impuestos a nivel póliza
  AD  Asociación documento     -> liga la póliza completa al UUID

Todas estas filas solo se generan cuando la póliza SÍ tiene un CFDI
conciliado (ver cfdi/cfdi_matcher.py). Si el movimiento no tiene
factura asociada (numero_factura vacío), se exporta igual que antes:
solo P + M1, sin AM/I/W2/V/AD.

El archivo de salida es **.xls real** (formato Excel 97-2003 / BIFF8),
que es lo que Contpaqi importa; ver exporters/xls_writer.py para cómo
se genera sin depender de paquetes que no están disponibles en este
entorno.

Los valores de catálogo específicos de Contpaqi que NO se pueden
deducir del CFDI ni del movimiento (ConceptoIVA, SubconceptoIVA,
ClasificadorIVA, Serie interna) se dejan como parámetros configurables
con el mismo valor que traía el archivo de referencia del usuario por
default; cada empresa debería confirmarlos contra su propio catálogo
de Contpaqi la primera vez.
"""

import os
import uuid as uuid_lib
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook

from xls_writer import guardar_como_xls, ErrorConversionXls

_RUTA_PLANTILLA_ENCABEZADOS = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "database",
    "plantilla_encabezados_contpaqi.xlsx",
)


@dataclass
class ImpuestoPoliza:
    base: float
    importe: float
    tasa: float  # ej. 0.16
    es_retencion: bool = False
    tipo_impuesto: str = "IVA"  # "IVA" | "ISR"; solo relevante si es_retencion=True


@dataclass
class FacturaAplicada:
    """Una factura (CFDI) conciliada contra un movimiento. Un movimiento
    puede traer más de una (pago que junta varias facturas, referencia
    'VARIOS' en Contpaqi) — ver MovimientoPoliza.facturas."""
    uuid: str
    serie: str = ""
    folio: str = ""
    rfc_persona: str = ""  # RFC del proveedor/cliente de ESTA factura
    impuestos: list = field(default_factory=list)  # list[ImpuestoPoliza]


@dataclass
class MovimientoPoliza:
    """Insumo para exportar: una póliza completa ya generada y cuadrada."""
    numero_poliza: int
    tipo: str            # 'ingreso' | 'egreso'
    fecha: datetime
    descripcion: str
    lineas: list          # lista de LineaPoliza (de policy_generator.py)
    tiene_iva: bool = False
    numero_factura: str = ""  # referencia real del movimiento; "" si no hay factura

    # --- Asociación de CFDI (opcional; la llena cfdi_matcher.py) ---
    # Un solo CFDI: usa cfdi_uuid/cfdi_serie/cfdi_folio/cfdi_impuestos.
    # Varios CFDI combinados en un mismo pago: usa `facturas` en su lugar
    # (el exportador prioriza `facturas` si trae algo).
    cfdi_uuid: Optional[str] = None
    cfdi_serie: Optional[str] = None
    cfdi_folio: Optional[str] = None
    cfdi_impuestos: list = field(default_factory=list)  # list[ImpuestoPoliza]
    facturas: list = field(default_factory=list)  # list[FacturaAplicada]
    cuenta_banco: Optional[str] = None   # requerido si cfdi_uuid está presente
    ret_iva: float = 0.0
    ret_isr: float = 0.0
    id_persona: str = ""  # código de proveedor/cliente en el catálogo de Contpaqi


@dataclass
class ConfiguracionCatalogoImpuestos:
    """Valores de catálogo de Contpaqi para las filas 'I' (F4) que no
    vienen ni del CFDI ni del movimiento. Cada empresa los confirma una
    vez contra su propio Contpaqi; el default es el que traía el
    archivo de referencia proporcionado."""
    concepto_iva: str = "201"
    subconcepto_iva: str = "201.03"
    clasificador_iva: str = "0"
    proporcion_diot: int = 100
    deducible_diot: int = 1
    origen: int = 2
    impuesto: str = "2"           # catálogo interno de Contpaqi para IVA
    objeto_impuesto: str = "2"


def _nuevo_guid() -> str:
    return str(uuid_lib.uuid4()).upper()


def _fila_p(mov: MovimientoPoliza) -> list:
    xml_tag = " (ADJUNTAR XML)" if mov.tiene_iva else ""
    return [
        "P", mov.fecha, 1 if mov.tipo == "ingreso" else 2, mov.numero_poliza,
        1, "0", f"{mov.descripcion}{xml_tag}", 11, 0, 0, _nuevo_guid(),
    ]


def _fila_m1(mov: MovimientoPoliza, linea, referencia: str) -> list:
    cuenta_texto = str(linea.cuenta).replace("-", "")
    return [
        "M1", cuenta_texto, referencia,
        0 if linea.naturaleza == "cargo" else 1,
        abs(round(linea.importe, 2)), "0", 0, linea.descripcion, " ",
        _nuevo_guid(), mov.fecha,
    ]


def _fila_am(uuid_cfdi: str) -> list:
    return ["AM", uuid_cfdi]


def _fila_i(mov: MovimientoPoliza, impuesto: ImpuestoPoliza,
            cat: ConfiguracionCatalogoImpuestos) -> list:
    total_impuesto = round(impuesto.base + impuesto.importe, 2)
    return [
        "I",
        mov.id_persona or "",
        mov.fecha.year, mov.fecha.month,
        mov.cuenta_banco, 1,
        mov.cfdi_serie or "", mov.cfdi_folio or "", None,
        mov.cfdi_uuid,
        cat.origen, 1, 1, 1,
        cat.impuesto, cat.objeto_impuesto, None,
        round(impuesto.tasa * 100, 4),
        round(impuesto.base, 2), round(impuesto.importe, 2), total_impuesto,
        1, 0, 0, None,
        round(impuesto.base, 2), 0, 0,
        _nuevo_guid(), None, 0,
        cat.concepto_iva, cat.subconcepto_iva, cat.clasificador_iva,
        cat.proporcion_diot, cat.deducible_diot,
    ]


def _fila_w2(impuesto: ImpuestoPoliza) -> list:
    return ["W2", round(impuesto.base, 2), 0, 0, None]


def _fila_v(mov: MovimientoPoliza, cat: ConfiguracionCatalogoImpuestos) -> list:
    base_total = round(sum(i.base for i in mov.cfdi_impuestos), 2)
    iva_total = round(sum(i.importe for i in mov.cfdi_impuestos), 2)
    tasa_principal = mov.cfdi_impuestos[0].tasa if mov.cfdi_impuestos else 0.0
    total_factura = round(base_total + iva_total, 2)
    return [
        "V",
        mov.id_persona or "",
        total_factura, round(tasa_principal * 100, 4), base_total, iva_total,
        1, 0,
        mov.cfdi_serie or "", mov.cfdi_folio or "", None, 0,
        round(total_factura - mov.ret_iva - mov.ret_isr, 2),
        mov.ret_iva, mov.ret_isr, total_factura,
        mov.fecha.year, mov.fecha.month, mov.cuenta_banco, 0,
        mov.cfdi_uuid, None, 0,
    ]


def _fila_ad(uuid_cfdi: str) -> list:
    return ["AD", uuid_cfdi]


def _construir_filas(movimientos_poliza: list, cat: ConfiguracionCatalogoImpuestos):
    filas = []
    con_cfdi = 0
    sin_cfdi = 0

    for mov in movimientos_poliza:
        filas.append(_fila_p(mov))
        referencia = (mov.numero_factura or "").strip()

        for linea in mov.lineas:
            filas.append(_fila_m1(mov, linea, referencia))
            if mov.cfdi_uuid:
                filas.append(_fila_am(mov.cfdi_uuid))

        if mov.cfdi_uuid and mov.cfdi_impuestos:
            if not mov.cuenta_banco:
                raise ValueError(
                    f"La póliza #{mov.numero_poliza} tiene CFDI asociado "
                    f"({mov.cfdi_uuid}) pero no trae cuenta_banco, necesaria "
                    f"para las filas 'I'/'V' (F4)."
                )
            for impuesto in mov.cfdi_impuestos:
                filas.append(_fila_i(mov, impuesto, cat))
                filas.append(_fila_w2(impuesto))
            filas.append(_fila_v(mov, cat))
            filas.append(_fila_ad(mov.cfdi_uuid))
            con_cfdi += 1
        else:
            sin_cfdi += 1

    return filas, con_cfdi, sin_cfdi


def exportar_polizas_contpaqi(
    movimientos_poliza: list[MovimientoPoliza],
    ruta_salida: str,
    catalogo_impuestos: Optional[ConfiguracionCatalogoImpuestos] = None,
) -> dict:
    """
    Genera el archivo .xls final (formato real de Contpaqi, con el
    bloque de encabezados de esquema al inicio) a partir de una lista
    de pólizas ya generadas y cuadradas.

    `ruta_salida` debe terminar en .xls; si no, se corrige solo.
    """
    if not ruta_salida.lower().endswith(".xls"):
        ruta_salida = os.path.splitext(ruta_salida)[0] + ".xls"

    cat = catalogo_impuestos or ConfiguracionCatalogoImpuestos()

    wb_plantilla = load_workbook(_RUTA_PLANTILLA_ENCABEZADOS)
    ws_plantilla = wb_plantilla["Datos"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Copiar el bloque de encabezados/esquema (filas 1-22) tal cual lo
    # espera Contpaqi, tomado del archivo de referencia real.
    for fila in ws_plantilla.iter_rows(values_only=True):
        ws.append(list(fila))

    filas_datos, con_cfdi, sin_cfdi = _construir_filas(movimientos_poliza, cat)
    for fila in filas_datos:
        ws.append(fila)

    try:
        ruta_final = guardar_como_xls(wb, ruta_salida)
        formato_final = "xls"
    except (ErrorConversionXls, FileNotFoundError) as e:
        ruta_final = os.path.splitext(ruta_salida)[0] + ".xlsx"
        wb.save(ruta_final)
        formato_final = f"xlsx (no se pudo generar .xls real: {e})"

    return {
        "archivo": ruta_final,
        "formato": formato_final,
        "polizas": len(movimientos_poliza),
        "polizas_con_cfdi": con_cfdi,
        "polizas_sin_cfdi": sin_cfdi,
        "filas_totales_datos": len(filas_datos),
    }
